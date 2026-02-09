#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Automated Economic Dashboard Data Updater
Fetches latest economic indicators and market data from public APIs
"""
import json
import os
import sys
from datetime import datetime, timedelta
import urllib.request
import urllib.error
import urllib.parse
import time
import tempfile

# API endpoints
# Check environment variable first, then fall back to local config file
FRED_API_KEY = os.environ.get('FRED_API_KEY', '')
EIA_API_KEY = os.environ.get('EIA_API_KEY', '')

# Try reading from local config file (for local development)
config_path = os.path.join(os.path.dirname(__file__), '.api_keys')
if os.path.exists(config_path):
    with open(config_path) as f:
        for line in f:
            line = line.strip()
            if line.startswith('FRED_API_KEY=') and not FRED_API_KEY:
                FRED_API_KEY = line.split('=', 1)[1].strip()
            elif line.startswith('EIA_API_KEY=') and not EIA_API_KEY:
                EIA_API_KEY = line.split('=', 1)[1].strip()

def print_safe(msg):
    """Print with safe encoding for Windows"""
    try:
        print(msg)
    except:
        print(msg.encode('ascii', 'replace').decode('ascii'))

def fetch_json(url, error_msg="Error fetching data"):
    """Fetch JSON data from URL with error handling"""
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=30) as response:
            return json.loads(response.read().decode())
    except urllib.error.HTTPError as e:
        print_safe(f"{error_msg}: HTTP {e.code}")
        return None
    except Exception as e:
        print_safe(f"{error_msg}: {e}")
        return None

def fetch_gold_price():
    """Fetch daily gold prices from FreeGoldAPI.com (USD per troy ounce)"""
    url = "https://freegoldapi.com/data/latest.json"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=30) as response:
            data = json.loads(response.read().decode())

        # Get the most recent ~260 trading days of data
        # Data is already sorted by date ascending
        recent_data = data[-260:] if len(data) > 260 else data

        # Convert to our format
        gold_prices = []
        for record in recent_data:
            if 'date' in record and 'price' in record:
                try:
                    gold_prices.append({
                        'date': record['date'],
                        'value': float(record['price'])
                    })
                except (ValueError, TypeError):
                    continue

        return gold_prices
    except Exception as e:
        print_safe(f"Error fetching gold price: {e}")
        return None

def fetch_fred_data(series_id, limit=12):
    """Fetch data from Federal Reserve Economic Data (FRED)"""
    if not FRED_API_KEY:
        print_safe(f"  ! Skipping {series_id} - No FRED API key")
        return []

    url = f"https://api.stlouisfed.org/fred/series/observations"
    params = f"?series_id={series_id}&api_key={FRED_API_KEY}&file_type=json&sort_order=desc&limit={limit}"

    data = fetch_json(url + params, f"Error fetching {series_id}")
    if not data or 'observations' not in data:
        return []

    observations = []
    for obs in reversed(data['observations']):
        if obs['value'] != '.':
            observations.append({
                'date': obs['date'],
                'series_id': series_id,
                'value': float(obs['value']),
                'realtime_start': obs['realtime_start'],
                'realtime_end': obs['realtime_end']
            })

    return observations

def fetch_eia_electricity(state_id, sector='ALL', limit=24):
    """
    Fetch electricity retail price data from EIA API v2

    Args:
        state_id: Two-letter state code (e.g., 'CO', 'VT', 'CA') or 'US' for national
        sector: 'RES' (residential), 'COM' (commercial), 'IND' (industrial), 'ALL' (total)
        limit: Number of monthly observations to fetch

    Returns:
        List of {date, value} dicts with price in cents/kWh
    """
    if not EIA_API_KEY:
        print_safe(f"  ! Skipping EIA {state_id} - No EIA API key")
        return []

    # EIA API v2 endpoint for electricity retail sales
    base_url = "https://api.eia.gov/v2/electricity/retail-sales/data/"

    # Build query parameters
    params = {
        'api_key': EIA_API_KEY,
        'frequency': 'monthly',
        'data[0]': 'price',  # Average retail price (cents/kWh)
        'facets[stateid][]': state_id,
        'facets[sectorid][]': sector,
        'sort[0][column]': 'period',
        'sort[0][direction]': 'desc',
        'length': str(limit)
    }

    # Build URL with parameters
    param_str = '&'.join(f"{k}={v}" for k, v in params.items())
    url = f"{base_url}?{param_str}"

    data = fetch_json(url, f"Error fetching EIA {state_id}")
    if not data or 'response' not in data or 'data' not in data['response']:
        return []

    observations = []
    for record in reversed(data['response']['data']):
        if record.get('price') is not None:
            # Convert period (YYYY-MM) to date format (YYYY-MM-01)
            period = record.get('period', '')
            if len(period) == 7:  # YYYY-MM format
                date_str = f"{period}-01"
            else:
                date_str = period

            observations.append({
                'date': date_str,
                'value': float(record['price']),  # Already in cents/kWh
                'state': record.get('stateid', state_id),
                'sector': record.get('sectorid', sector)
            })

    return observations

def fetch_news_sentiment():
    """Fetch Daily News Sentiment Index from SF Fed Excel file"""
    # The URL pattern includes a cache-busting date parameter
    url = "https://www.frbsf.org/wp-content/uploads/news_sentiment_data.xlsx"

    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=60) as response:
            # Save to temp file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                tmp.write(response.read())
                tmp_path = tmp.name

        # Try to parse Excel file using openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path, read_only=True, data_only=True)
            ws = wb.active

            # Read data - typically date in column A, sentiment in column B
            observations = []
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[0] is None:
                    continue
                date_val = row[0]
                sentiment_val = row[1] if len(row) > 1 else None

                if sentiment_val is None:
                    continue

                # Handle date formatting
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_val)

                try:
                    observations.append({
                        'date': date_str,
                        'value': float(sentiment_val)
                    })
                except (ValueError, TypeError):
                    continue

            wb.close()
            os.unlink(tmp_path)

            # Return last 365 days of data for sparkline
            if observations:
                # Sort by date and get recent data
                observations.sort(key=lambda x: x['date'])
                return observations[-365:]
            return []

        except ImportError:
            print_safe("  ! openpyxl not installed, skipping News Sentiment")
            os.unlink(tmp_path)
            return []

    except Exception as e:
        print_safe(f"  ! News Sentiment unavailable: {e}")
        return []


def fetch_market_data_yahoo(symbol, historical=False):
    """Fetch market data from Yahoo Finance using chart API (more reliable)"""
    end_date = datetime.now()
    start_date = end_date - timedelta(days=365 if historical else 7)

    period1 = int(start_date.timestamp())
    period2 = int(end_date.timestamp())

    # Use the v8 chart API which is more reliable than the download API
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
    url += f"?period1={period1}&period2={period2}&interval=1d"

    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        with urllib.request.urlopen(req, timeout=30) as response:
            data = json.loads(response.read().decode('utf-8'))

            result = data.get('chart', {}).get('result', [])
            if not result:
                return None

            chart_data = result[0]
            timestamps = chart_data.get('timestamp', [])
            indicators = chart_data.get('indicators', {}).get('quote', [{}])[0]
            closes = indicators.get('close', [])

            if not timestamps or not closes:
                return None

            if historical:
                # Return all data points for sparkline
                history = []
                for i, ts in enumerate(timestamps):
                    if closes[i] is not None:
                        date_str = datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
                        history.append({
                            'date': date_str,
                            'close': closes[i]
                        })
                if not history:
                    return None
                return {
                    'symbol': symbol,
                    'current_date': history[-1]['date'] if history else None,
                    'current_close': history[-1]['close'] if history else None,
                    'history': history
                }
            else:
                # Just return latest value
                last_idx = len(closes) - 1
                while last_idx >= 0 and closes[last_idx] is None:
                    last_idx -= 1
                if last_idx < 0:
                    return None
                return {
                    'symbol': symbol,
                    'date': datetime.fromtimestamp(timestamps[last_idx]).strftime('%Y-%m-%d'),
                    'close': closes[last_idx],
                    'volume': 0
                }
    except Exception as e:
        print_safe(f"  ! {symbol} unavailable")
        return None

# Port-to-resort mapping for ski context
PORT_RESORT_MAPPING = {
    'Sweetgrass': {'resorts': 'Big Sky, Bridger Bowl', 'route': 'I-15 from Calgary/Edmonton'},
    'Roosville': {'resorts': 'Whitefish, Big Sky', 'route': 'US-93 from BC'},
    'Piegan': {'resorts': 'Glacier NP, Whitefish', 'route': 'US-89 from Alberta'},
    'Derby Line': {'resorts': 'Stowe, Jay Peak, Burke', 'route': 'I-91 from Montreal/Quebec'},
    'Highgate Springs': {'resorts': 'Stowe, Smugglers Notch', 'route': 'I-89 from Montreal'},
    'Blaine': {'resorts': 'Mt. Baker, Stevens Pass, Crystal', 'route': 'I-5 from Vancouver'},
    'Sumas': {'resorts': 'Mt. Baker', 'route': 'WA-9 from BC'},
    'Champlain': {'resorts': 'Whiteface, Gore, Lake Placid', 'route': 'I-87 from Montreal'},
    'Houlton': {'resorts': 'Sugarloaf, Sunday River', 'route': 'I-95 from New Brunswick'},
    'Calais': {'resorts': 'Sugarloaf, Sunday River', 'route': 'US-1 from New Brunswick'},
}


def fetch_bts_border_crossings(state=None, limit=36):
    """
    Fetch US/Canada border crossing data from Bureau of Transportation Statistics (BTS)

    Uses the Socrata API endpoint (no authentication required)
    Data source: https://data.bts.gov/resource/keg4-3bc2.json

    Args:
        state: State filter (e.g., 'Montana', 'Vermont') or None for all states
        limit: Number of months of data to return per port

    Returns:
        Dict with national totals, regional data, port-level details, and 2019 baseline
    """
    base_url = "https://data.bts.gov/resource/keg4-3bc2.json"

    # Calculate date range - include 2019 baseline for COVID comparison
    end_year = datetime.now().year
    start_year = 2019  # Include 2019 for baseline comparison

    # Query parameters - get Canadian border only, tourist-relevant measures
    # Personal Vehicle Passengers is the key metric for tourism
    params = {
        '$limit': 100000,
        '$where': f"border='US-Canada Border' AND date >= '{start_year}-01-01'",
        '$order': 'date DESC'
    }

    param_str = '&'.join(f"{k}={urllib.parse.quote(str(v))}" for k, v in params.items())
    url = f"{base_url}?{param_str}"

    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=120) as response:
            data = json.loads(response.read().decode())

        if not data:
            return None

        # Process data into structured format
        # Group by date for national totals
        from collections import defaultdict

        national_monthly = defaultdict(lambda: {'passengers': 0, 'vehicles': 0})
        state_monthly = defaultdict(lambda: defaultdict(lambda: {'passengers': 0, 'vehicles': 0}))
        port_monthly = defaultdict(lambda: defaultdict(lambda: {'passengers': 0, 'vehicles': 0}))

        # Key ski-relevant states (removed NH - no significant ski-relevant border ports)
        ski_states = ['Montana', 'Vermont', 'Washington', 'New York', 'Maine', 'Michigan', 'Minnesota', 'Idaho']

        for record in data:
            date_str = record.get('date', '')[:7]  # YYYY-MM format
            if not date_str:
                continue

            state_name = record.get('state', '')
            port_name = record.get('port_name', '')
            measure = record.get('measure', '')
            value = int(record.get('value', 0) or 0)

            # Aggregate national totals
            if measure == 'Personal Vehicle Passengers':
                national_monthly[date_str]['passengers'] += value
            elif measure == 'Personal Vehicles':
                national_monthly[date_str]['vehicles'] += value

            # State-level aggregation for ski states
            if state_name in ski_states:
                if measure == 'Personal Vehicle Passengers':
                    state_monthly[state_name][date_str]['passengers'] += value
                elif measure == 'Personal Vehicles':
                    state_monthly[state_name][date_str]['vehicles'] += value

                # Port-level for key ports
                if measure == 'Personal Vehicle Passengers':
                    port_monthly[port_name][date_str]['passengers'] += value

        # Convert to list format sorted by date
        def to_time_series(monthly_dict, limit=36):
            sorted_dates = sorted(monthly_dict.keys(), reverse=True)[:limit]
            return [
                {
                    'date': f"{d}-01",
                    'value': monthly_dict[d]['passengers'],
                    'vehicles': monthly_dict[d]['vehicles']
                }
                for d in reversed(sorted_dates)
            ]

        # Calculate 2019 baseline for comparison
        def calc_2019_baseline(monthly_dict):
            """Calculate average monthly value for 2019 (pre-COVID baseline)"""
            months_2019 = [m for m in monthly_dict.keys() if m.startswith('2019')]
            if months_2019:
                total = sum(monthly_dict[m]['passengers'] for m in months_2019)
                return int(total / len(months_2019))
            return None

        def calc_vs_2019(monthly_dict):
            """Calculate current vs 2019 baseline percentage"""
            baseline = calc_2019_baseline(monthly_dict)
            if not baseline:
                return None
            # Get most recent 12 months average
            recent_dates = sorted(monthly_dict.keys(), reverse=True)[:12]
            if len(recent_dates) < 6:
                return None
            recent_avg = sum(monthly_dict[d]['passengers'] for d in recent_dates) / len(recent_dates)
            return round((recent_avg - baseline) / baseline * 100, 1)

        # National totals
        national_series = to_time_series(national_monthly, limit)
        national_baseline_2019 = calc_2019_baseline(national_monthly)
        national_vs_2019 = calc_vs_2019(national_monthly)

        # State series for ski-relevant states
        state_series = {}
        state_baselines = {}
        for state_name in ski_states:
            if state_name in state_monthly and state_monthly[state_name]:
                series = to_time_series(state_monthly[state_name], limit)
                if series:
                    key = state_name.lower().replace(' ', '_')
                    state_series[key] = series
                    vs_2019 = calc_vs_2019(state_monthly[state_name])
                    if vs_2019 is not None:
                        state_baselines[key] = vs_2019

        # Key ports (top ports by volume in ski-relevant states)
        key_ports = list(PORT_RESORT_MAPPING.keys())

        port_series = {}
        for port_name in key_ports:
            if port_name in port_monthly and port_monthly[port_name]:
                series = to_time_series(port_monthly[port_name], limit)
                if series:
                    # Clean port name for JSON key
                    key = port_name.lower().replace(' ', '_').replace('-', '_')
                    port_info = PORT_RESORT_MAPPING.get(port_name, {})
                    port_series[key] = {
                        'data': series,
                        'resorts': port_info.get('resorts', ''),
                        'route': port_info.get('route', ''),
                        'vs_2019': calc_vs_2019(port_monthly[port_name])
                    }

        return {
            'national': national_series,
            'national_baseline_2019': national_baseline_2019,
            'national_vs_2019': national_vs_2019,
            'states': state_series,
            'state_baselines': state_baselines,
            'ports': port_series
        }

    except Exception as e:
        print_safe(f"  ! BTS border crossing data unavailable: {e}")
        return None


def fetch_statcan_canadian_outbound():
    """
    Fetch Canadian outbound travel to US by province from Statistics Canada.

    Uses Table 24-10-0071-01: Visits, nights and expenditures for Canadian
    residents travelling in Canada and abroad.

    Returns:
        Dict with quarterly visits to US by province of origin, with 2019 baseline
    """
    import csv
    import zipfile
    import io
    from collections import defaultdict

    csv_url = "https://www150.statcan.gc.ca/n1/tbl/csv/24100071-eng.zip"

    try:
        req = urllib.request.Request(csv_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=120) as response:
            zip_data = io.BytesIO(response.read())

        with zipfile.ZipFile(zip_data) as zf:
            csv_filename = [n for n in zf.namelist() if n.endswith('.csv') and 'MetaData' not in n][0]
            with zf.open(csv_filename) as f:
                content = io.TextIOWrapper(f, encoding='utf-8-sig')
                reader = csv.DictReader(content)

                # Key provinces for ski market analysis
                target_provinces = [
                    'Alberta (trip origin)',      # Montana, Idaho feeders
                    'British Columbia (trip origin)',  # Washington, Montana feeders
                    'Ontario (trip origin)',      # Vermont, NY feeders
                    'Quebec (trip origin)',       # Vermont, Maine feeders
                    'Canada (trip origin)'        # National total
                ]

                data = defaultdict(lambda: defaultdict(float))

                for row in reader:
                    # Filter: US destination, key provinces, Visits metric, All durations
                    if (row.get('GEO') == 'United States of America' and
                        row.get('Province of trip origin') in target_provinces and
                        row.get('Visit duration') == 'All visit durations' and
                        row.get('Visit statistics') == 'Visits' and
                        row.get('VALUE')):

                        date = row['REF_DATE']
                        province = row['Province of trip origin'].replace(' (trip origin)', '')
                        value = float(row['VALUE'])  # in thousands
                        data[province][date] = value

        # Build output structure
        result = {}
        province_mapping = {
            'Canada': {'key': 'canada_total', 'feeder_for': 'All US ski markets'},
            'Alberta': {'key': 'alberta', 'feeder_for': 'Montana, Idaho, Utah'},
            'British Columbia': {'key': 'british_columbia', 'feeder_for': 'Washington, Montana, Colorado'},
            'Ontario': {'key': 'ontario', 'feeder_for': 'Vermont, New York, Michigan'},
            'Quebec': {'key': 'quebec', 'feeder_for': 'Vermont, Maine, New Hampshire'}
        }

        for prov_name, prov_info in province_mapping.items():
            prov_data = data.get(prov_name, {})
            if not prov_data:
                continue

            # Get sorted dates (most recent first)
            sorted_dates = sorted(prov_data.keys(), reverse=True)

            # Calculate 2019 baseline average
            q2019 = [prov_data.get(f'2019-{m:02d}', 0) for m in [1, 4, 7, 10]]
            valid_2019 = [q for q in q2019 if q > 0]
            baseline_2019 = sum(valid_2019) / len(valid_2019) if valid_2019 else None

            # Latest value and vs 2019
            latest_date = sorted_dates[0] if sorted_dates else None
            latest_value = prov_data.get(latest_date, 0) if latest_date else 0
            vs_2019 = round((latest_value - baseline_2019) / baseline_2019 * 100, 1) if baseline_2019 else None

            # Build time series (last 12 quarters)
            series = [
                {'date': d, 'value': int(prov_data[d])}
                for d in reversed(sorted_dates[:12])
            ]

            result[prov_info['key']] = {
                'name': prov_name,
                'feeder_for': prov_info['feeder_for'],
                'data': series,
                'latest': int(latest_value),
                'latest_date': latest_date,
                'baseline_2019': int(baseline_2019) if baseline_2019 else None,
                'vs_2019': vs_2019
            }

        return result

    except Exception as e:
        print_safe(f"  ! StatCan Canadian outbound data unavailable: {e}")
        return None


def fetch_tsa_checkpoint_data(limit=90):
    """
    Fetch TSA checkpoint travel numbers from TSA.gov

    Scrapes the daily passenger volume data from TSA's public pages.
    Data is updated weekdays by 9am with ~1 day lag.

    Args:
        limit: Number of days of data to return

    Returns:
        List of dicts with date and value (passengers screened)
    """
    import re
    from datetime import datetime

    results = []
    current_year = datetime.now().year

    # Fetch current year and previous year if needed
    years_to_fetch = [current_year, current_year - 1]

    for year in years_to_fetch:
        if year == current_year:
            url = "https://www.tsa.gov/travel/passenger-volumes"
        else:
            url = f"https://www.tsa.gov/travel/passenger-volumes/{year}"

        try:
            req = urllib.request.Request(url, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            })
            with urllib.request.urlopen(req, timeout=60) as response:
                html = response.read().decode('utf-8')

            # TSA table structure: rows with date in first cell, numbers in subsequent cells
            # Extract all table rows with date patterns
            # Pattern: <tr>...<td>M/D/YYYY</td><td>NUMBER</td>...</tr>
            row_pattern = r'<tr[^>]*>.*?<td[^>]*>(\d{1,2}/\d{1,2}/\d{4})</td>\s*<td[^>]*>([\d,]+)</td>'
            matches = re.findall(row_pattern, html, re.DOTALL)

            for date_str, value_str in matches:
                try:
                    # Parse date M/D/YYYY to YYYY-MM-DD
                    parts = date_str.split('/')
                    month, day, yr = int(parts[0]), int(parts[1]), int(parts[2])
                    iso_date = f"{yr}-{month:02d}-{day:02d}"

                    # Parse value (remove commas)
                    value = int(value_str.replace(',', ''))

                    results.append({
                        'date': iso_date,
                        'value': value
                    })
                except (ValueError, IndexError):
                    continue

        except Exception as e:
            # Continue with other years even if one fails
            continue

    if not results:
        print_safe("  ! TSA checkpoint data: No data parsed")
        return None

    # Remove duplicates (in case of overlapping data)
    seen = set()
    unique_results = []
    for r in results:
        if r['date'] not in seen:
            seen.add(r['date'])
            unique_results.append(r)

    # Sort by date ascending
    unique_results.sort(key=lambda x: x['date'])

    # Return most recent 'limit' entries
    return unique_results[-limit:] if unique_results else None


def fetch_bts_t100_aggregate(limit=36):
    """
    Fetch aggregate T-100 air traffic data from BTS

    Uses the preliminary estimates dataset which provides monthly
    national-level passenger counts.

    Args:
        limit: Number of months to return

    Returns:
        List of dicts with date, passengers (segment), and checks (TSA comparison)
    """
    url = "https://data.bts.gov/resource/3xj5-daif.json?$order=date%20DESC&$limit=100"

    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=60) as response:
            data = json.loads(response.read().decode())

        if not data:
            return None

        results = []
        for record in data[:limit]:
            date_str = record.get('date', '')[:10]  # YYYY-MM-DD
            if not date_str:
                continue

            # segment = total passengers transported
            # market = enplanements (unduplicated)
            # checks = TSA screenings
            # Use _fit values as fallback for recent months that don't have final data
            segment = int(float(record.get('segment') or record.get('segment_fit') or 0))
            market = int(float(record.get('market') or record.get('market_fit') or 0))
            checks = int(float(record.get('checks', 0) or 0))

            results.append({
                'date': date_str,
                'value': segment,  # Primary metric: passengers transported
                'enplanements': market,
                'tsa_checks': checks
            })

        # Sort ascending by date
        results.sort(key=lambda x: x['date'])
        return results[-limit:] if results else None

    except Exception as e:
        print_safe(f"  ! BTS T-100 aggregate data unavailable: {e}")
        return None


# Ski gateway airport configuration
SKI_GATEWAY_AIRPORTS = {
    # Colorado
    'HDN': {'name': 'Yampa Valley Regional', 'city': 'Hayden', 'resorts': 'Steamboat Springs', 'region': 'colorado'},
    'EGE': {'name': 'Eagle County Regional', 'city': 'Eagle', 'resorts': 'Vail, Beaver Creek', 'region': 'colorado'},
    'ASE': {'name': 'Aspen-Pitkin County', 'city': 'Aspen', 'resorts': 'Aspen, Snowmass', 'region': 'colorado'},
    'MTJ': {'name': 'Montrose Regional', 'city': 'Montrose', 'resorts': 'Telluride', 'region': 'colorado'},
    'GUC': {'name': 'Gunnison-Crested Butte', 'city': 'Gunnison', 'resorts': 'Crested Butte', 'region': 'colorado'},
    'DRO': {'name': 'Durango-La Plata County', 'city': 'Durango', 'resorts': 'Purgatory', 'region': 'colorado'},
    'TEX': {'name': 'Telluride Regional', 'city': 'Telluride', 'resorts': 'Telluride', 'region': 'colorado'},
    # Northern Rockies (Montana, Wyoming, Idaho)
    'JAC': {'name': 'Jackson Hole', 'city': 'Jackson', 'resorts': 'Jackson Hole, Grand Targhee', 'region': 'rockies'},
    'SUN': {'name': 'Friedman Memorial', 'city': 'Hailey', 'resorts': 'Sun Valley', 'region': 'rockies'},
    'FCA': {'name': 'Glacier Park Intl', 'city': 'Kalispell', 'resorts': 'Whitefish Mountain', 'region': 'rockies'},
    'BZN': {'name': 'Bozeman Yellowstone', 'city': 'Bozeman', 'resorts': 'Big Sky, Bridger Bowl', 'region': 'rockies'},
    'MSO': {'name': 'Missoula Intl', 'city': 'Missoula', 'resorts': 'Snowbowl, Lost Trail', 'region': 'rockies'},
    # California/Nevada
    'RNO': {'name': 'Reno-Tahoe Intl', 'city': 'Reno', 'resorts': 'Lake Tahoe resorts', 'region': 'california'},
    'MMH': {'name': 'Mammoth Yosemite', 'city': 'Mammoth Lakes', 'resorts': 'Mammoth Mountain', 'region': 'california'},
    'FAT': {'name': 'Fresno Yosemite Intl', 'city': 'Fresno', 'resorts': 'Yosemite, China Peak', 'region': 'california'},
    'PSP': {'name': 'Palm Springs Intl', 'city': 'Palm Springs', 'resorts': 'Big Bear, Mountain High', 'region': 'california'},
    # Major Hubs (high volume, serve multiple resorts)
    'DEN': {'name': 'Denver Intl', 'city': 'Denver', 'resorts': 'All Colorado resorts', 'region': 'hubs'},
    'SLC': {'name': 'Salt Lake City Intl', 'city': 'Salt Lake City', 'resorts': 'Park City, Snowbird, Alta', 'region': 'hubs'},
    # Canadian ski gateways (T-100 International data)
    'YYC': {'name': 'Calgary Intl', 'city': 'Calgary', 'resorts': 'Banff, Lake Louise, Kicking Horse', 'region': 'canada'},
    'YVR': {'name': 'Vancouver Intl', 'city': 'Vancouver', 'resorts': 'Whistler Blackcomb', 'region': 'canada'},
    'YLW': {'name': 'Kelowna Intl', 'city': 'Kelowna', 'resorts': 'Big White, Silver Star', 'region': 'canada'},
    'YXC': {'name': 'Cranbrook/Kimberley', 'city': 'Cranbrook', 'resorts': 'Fernie, Kimberley', 'region': 'canada'},
    'YEG': {'name': 'Edmonton Intl', 'city': 'Edmonton', 'resorts': 'Marmot Basin, Jasper', 'region': 'canada'},
}


def fetch_ski_gateway_airports():
    """
    Load ski gateway airport data with true month-over-month YoY comparisons.

    Uses pre-processed monthly data from airport_passengers.json (generated by
    generate_airport_output.py from BTS T-100 CSV downloads and direct airport scrapers).

    Returns:
        List of dicts with airport code, latest month passengers, and real YoY change
    """
    airport_json_path = os.path.join(os.path.dirname(__file__), 'static', 'data', 'airport_passengers.json')

    try:
        with open(airport_json_path, 'r') as f:
            airport_data = json.load(f)

        if not airport_data or 'airports' not in airport_data:
            print_safe("  ! airport_passengers.json missing or invalid")
            return None

        results = []
        months_abbrev = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        for code, monthly_data in airport_data['airports'].items():
            if code not in SKI_GATEWAY_AIRPORTS:
                continue

            # Get all months sorted descending
            sorted_months = sorted(monthly_data.keys(), reverse=True)
            if not sorted_months:
                continue

            # Latest month data
            latest_month = sorted_months[0]  # e.g., "2025-11"
            latest = monthly_data[latest_month]

            # Parse year and month
            year, month = latest_month.split('-')
            year = int(year)
            month = int(month)

            # Get prior year same month for true YoY comparison
            prior_month_key = f"{year - 1}-{month:02d}"
            yoy_change = None

            if prior_month_key in monthly_data:
                prior_pax = monthly_data[prior_month_key]['passengers']
                if prior_pax > 0:
                    yoy_change = ((latest['passengers'] - prior_pax) / prior_pax) * 100

            # Also use the pre-computed yoy_pct if available and our calculation failed
            if yoy_change is None and latest.get('yoy_pct') is not None:
                yoy_change = latest['yoy_pct']

            config = SKI_GATEWAY_AIRPORTS[code]
            results.append({
                'code': code,
                'name': config['name'],
                'city': config['city'],
                'resorts': config['resorts'],
                'region': config['region'],
                'month': f"{months_abbrev[month - 1]} {year}",
                'month_key': latest_month,
                'passengers': latest['passengers'],
                'yoy_change': round(yoy_change, 1) if yoy_change is not None else None,
                'data_source': airport_data.get('data_sources', {}).get(code, {}).get('source', 'BTS T-100'),
                'months_available': len(sorted_months)
            })

        # Sort by passengers descending
        results.sort(key=lambda x: x['passengers'], reverse=True)
        return results

    except FileNotFoundError:
        print_safe(f"  ! airport_passengers.json not found - run generate_airport_output.py first")
        return None
    except Exception as e:
        print_safe(f"  ! Ski gateway airport data unavailable: {e}")
        return None


def build_economic_narrative(dashboard_data):
    """Build plain-text economic context narrative for ski industry.

    Uses factual, measured language per CLAUDE.md guidelines.
    Returns None if insufficient data for a meaningful narrative.
    """
    parts = []

    # Consumer confidence trend
    cc = dashboard_data.get('consumer_confidence', [])
    if len(cc) >= 2:
        latest = cc[-1]['value']
        prior = cc[-2]['value']
        direction = "rose" if latest > prior else "fell" if latest < prior else "held steady"
        if latest < 60:
            parts.append(f"Consumer confidence {direction} to {latest:.1f}, in pessimistic territory that may reduce discretionary travel spending.")
        elif latest > 85:
            parts.append(f"Consumer confidence {direction} to {latest:.1f}, elevated territory that tends to support discretionary travel spending.")
        else:
            parts.append(f"Consumer confidence {direction} to {latest:.1f}.")

    # Inflation
    cpi_yoy = dashboard_data.get('cpi_yoy', [])
    if cpi_yoy:
        inflation = cpi_yoy[-1]['value']
        if inflation > 4:
            parts.append(f"Inflation at {inflation:.1f}% is well above the Fed's 2% target, squeezing consumer purchasing power and increasing resort operating costs.")
        elif inflation > 2.5:
            parts.append(f"Inflation at {inflation:.1f}% remains above the Fed's 2% target.")
        else:
            parts.append(f"Inflation at {inflation:.1f}% is near the Fed's 2% target.")

    # Savings rate
    savings = dashboard_data.get('personal_savings_rate', [])
    if savings:
        rate = savings[-1]['value']
        if rate < 4:
            parts.append(f"The personal savings rate ({rate:.1f}%) is low, suggesting consumers may have less buffer for discretionary spending like ski trips.")
        elif rate > 8:
            parts.append(f"The personal savings rate ({rate:.1f}%) is elevated, suggesting consumers may have a savings cushion for discretionary spending.")

    # Yield curve
    spread = dashboard_data.get('yield_curve_spread', [])
    if spread:
        latest_spread = spread[-1]['value']
        if latest_spread < 0:
            parts.append(f"The yield curve is inverted ({latest_spread:+.2f}%), a historically reliable recession indicator.")
        elif latest_spread < 0.3:
            parts.append(f"The yield curve is nearly flat ({latest_spread:+.2f}%), suggesting economic uncertainty.")

    return " ".join(parts) if parts else None


def update_dashboard():
    """Main function to update dashboard data"""
    print_safe("Starting dashboard update...")
    print_safe(f"Timestamp: {datetime.now().isoformat()}")

    dashboard_data = {
        'updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    # 1. Consumer Confidence
    print_safe("\nFetching Consumer Confidence...")
    consumer_confidence = fetch_fred_data('UMCSENT', limit=12)
    if consumer_confidence:
        dashboard_data['consumer_confidence'] = consumer_confidence
        print_safe(f"  OK Latest: {consumer_confidence[-1]['value']} ({consumer_confidence[-1]['date']})")

    # 2. Market Data - Using FRED S&P 500 Index
    print_safe("\nFetching Market Data (S&P 500 from FRED)...")
    sp500_data = fetch_fred_data('SP500', limit=12)
    if sp500_data:
        # Format to match expected structure
        markets = [{
            'symbol': 'SPY',
            'current_date': sp500_data[-1]['date'],
            'current_close': sp500_data[-1]['value'],
            'history': [{'date': d['date'], 'close': d['value']} for d in sp500_data]
        }]
        dashboard_data['markets'] = markets
        print_safe(f"  OK S&P 500: {sp500_data[-1]['value']:.2f} ({len(sp500_data)} data points)")

    # 2b. Dow Jones Industrial Average (Daily - fetch ~1 year of trading days)
    print_safe("\nFetching Dow Jones Industrial Average...")
    djia_data = fetch_fred_data('DJIA', limit=260)
    if djia_data:
        dashboard_data['djia'] = djia_data
        print_safe(f"  OK DJIA: {djia_data[-1]['value']:.2f} ({len(djia_data)} data points)")

    # 2c. NASDAQ Composite Index (Daily - fetch ~1 year of trading days)
    print_safe("\nFetching NASDAQ Composite Index...")
    nasdaq_data = fetch_fred_data('NASDAQCOM', limit=260)
    if nasdaq_data:
        dashboard_data['nasdaq'] = nasdaq_data
        print_safe(f"  OK NASDAQ: {nasdaq_data[-1]['value']:.2f} ({len(nasdaq_data)} data points)")

    # 2d. Hotel & Lodging REITs - use FRED NASDAQ REIT Index
    # NASDAQNQUSB351020 = Nasdaq US Benchmark Real Estate Investment Trusts Index
    print_safe("\nFetching Hotel & Lodging REITs (Nasdaq REIT Index)...")
    reit_data = fetch_fred_data('NASDAQNQUSB351020', limit=260)
    if reit_data:
        dashboard_data['hotel_reits'] = reit_data
        print_safe(f"  OK REITs: {reit_data[-1]['value']:.2f} ({len(reit_data)} data points)")

    # 2e. VIX Volatility Index (Daily - fetch ~1 year of trading days)
    print_safe("\nFetching VIX Volatility Index...")
    vix_data = fetch_fred_data('VIXCLS', limit=260)
    if vix_data:
        dashboard_data['vix'] = vix_data
        print_safe(f"  OK VIX: {vix_data[-1]['value']:.2f} ({len(vix_data)} data points)")

    # 2f. Magnificent 7 Stocks
    print_safe("\nFetching Magnificent 7 Stocks...")
    mag7_symbols = ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'NVDA', 'META', 'TSLA']
    mag7_names = {
        'AAPL': 'Apple Inc.',
        'MSFT': 'Microsoft Corp',
        'GOOGL': 'Alphabet Inc.',
        'AMZN': 'Amazon.com Inc.',
        'NVDA': 'NVIDIA Corp',
        'META': 'Meta Platforms',
        'TSLA': 'Tesla Inc.'
    }
    mag7_data = []
    for symbol in mag7_symbols:
        stock_data = fetch_market_data_yahoo(symbol, historical=True)
        if stock_data:
            # Transform to expected format
            mag7_data.append({
                'symbol': symbol,
                'name': mag7_names[symbol],
                'price': stock_data['current_close'],
                'previousClose': stock_data['history'][-2]['close'] if len(stock_data['history']) >= 2 else stock_data['current_close'],
                'history': [{'date': h['date'], 'value': h['close']} for h in stock_data['history'][-90:]]  # Last 90 days
            })
            print_safe(f"  OK {symbol}: ${stock_data['current_close']:.2f}")
        else:
            print_safe(f"  ! {symbol}: unavailable")
        time.sleep(0.5)  # Rate limit
    if mag7_data:
        dashboard_data['mag7'] = mag7_data
        print_safe(f"  Fetched {len(mag7_data)} of 7 stocks")

    # 3. Fed Funds Rate (Daily - fetch ~1 year of trading days)
    print_safe("\nFetching Fed Funds Rate...")
    fed_rate = fetch_fred_data('DFF', limit=260)
    if fed_rate:
        dashboard_data['fed_funds_rate'] = fed_rate
        print_safe(f"  OK Latest: {fed_rate[-1]['value']}% ({len(fed_rate)} data points)")

    # 4. Unemployment Rate
    print_safe("\nFetching Unemployment Rate...")
    unemployment = fetch_fred_data('UNRATE', limit=12)
    if unemployment:
        dashboard_data['unemployment'] = unemployment
        print_safe(f"  OK Latest: {unemployment[-1]['value']}% ({len(unemployment)} data points)")

    # 5. CPI (Consumer Price Index) - fetch 24 months for YoY calculation
    print_safe("\nFetching CPI...")
    cpi_data = fetch_fred_data('CPIAUCSL', limit=24)
    if cpi_data:
        dashboard_data['cpi'] = cpi_data
        print_safe(f"  OK Latest: {cpi_data[-1]['value']} ({len(cpi_data)} data points)")

        # Compute YoY inflation rate from CPI index
        cpi_yoy = []
        for i in range(len(cpi_data)):
            # Find a record ~12 months prior
            current = cpi_data[i]
            current_date = current['date'][:7]  # YYYY-MM
            for j in range(i):
                prior_date = cpi_data[j]['date'][:7]
                # Check if approximately 12 months apart
                cy, cm = int(current_date[:4]), int(current_date[5:7])
                py, pm = int(prior_date[:4]), int(prior_date[5:7])
                months_diff = (cy - py) * 12 + (cm - pm)
                if 11 <= months_diff <= 13:
                    prior = cpi_data[j]
                    if prior['value'] > 0:
                        rate = round(((current['value'] - prior['value']) / prior['value']) * 100, 1)
                        cpi_yoy.append({'date': current['date'], 'value': rate})
                    break
        if cpi_yoy:
            dashboard_data['cpi_yoy'] = cpi_yoy
            print_safe(f"  OK CPI YoY Inflation: {cpi_yoy[-1]['value']}% ({len(cpi_yoy)} data points)")

    # 6. Employment (Total Nonfarm Payroll)
    print_safe("\nFetching Employment...")
    employment_data = fetch_fred_data('PAYEMS', limit=12)
    if employment_data:
        dashboard_data['employment'] = employment_data
        print_safe(f"  OK Latest: {employment_data[-1]['value']} thousand ({len(employment_data)} data points)")

    # 7. Average Hourly Earnings (Wages)
    print_safe("\nFetching Wages...")
    wages_data = fetch_fred_data('CES0500000003', limit=12)
    if wages_data:
        dashboard_data['wages'] = wages_data
        print_safe(f"  OK Latest: ${wages_data[-1]['value']} ({len(wages_data)} data points)")

    # 8. GDP (Quarterly)
    print_safe("\nFetching GDP...")
    gdp_data = fetch_fred_data('GDP', limit=12)
    if gdp_data:
        dashboard_data['gdp'] = gdp_data
        print_safe(f"  OK Latest: ${gdp_data[-1]['value']} billion ({len(gdp_data)} data points)")

    # 9. Housing Starts (Monthly)
    print_safe("\nFetching Housing Starts...")
    housing_data = fetch_fred_data('HOUST', limit=12)
    if housing_data:
        dashboard_data['housing_starts'] = housing_data
        print_safe(f"  OK Latest: {housing_data[-1]['value']} thousand units ({len(housing_data)} data points)")

    # 9b. Bankruptcy Activity (PPI - Producer Price Index for Bankruptcy Legal Services)
    # PCU541110541110903 = PPI for Offices of Lawyers: Bankruptcy and Other Business Legal Services
    print_safe("\nFetching Bankruptcy Activity (PPI)...")
    bankruptcy_data = fetch_fred_data('PCU541110541110903', limit=12)
    if bankruptcy_data:
        dashboard_data['bankruptcy_ppi'] = bankruptcy_data
        print_safe(f"  OK Latest: {bankruptcy_data[-1]['value']} ({len(bankruptcy_data)} data points)")

    # 9c. Personal Savings Rate (Monthly)
    print_safe("\nFetching Personal Savings Rate...")
    personal_savings = fetch_fred_data('PSAVERT', limit=12)
    if personal_savings:
        dashboard_data['personal_savings_rate'] = personal_savings
        print_safe(f"  OK Latest: {personal_savings[-1]['value']}% ({len(personal_savings)} data points)")

    # 9d. PCE Recreation Spending (Monthly, Billions $)
    print_safe("\nFetching PCE Recreation Spending...")
    pce_recreation = fetch_fred_data('DPCERA3M086SBEA', limit=12)
    if pce_recreation:
        dashboard_data['pce_recreation'] = pce_recreation
        print_safe(f"  OK Latest: ${pce_recreation[-1]['value']}B ({len(pce_recreation)} data points)")

    # =========================================================================
    # ELECTRICITY PRICING INDICATORS (Ski Region Focus)
    # Using EIA API for state-level data (cents/kWh, commercial sector)
    # =========================================================================

    print_safe("\n--- Electricity Pricing (EIA State-Level, Commercial Sector) ---")

    # Key ski states with their regions:
    # - Colorado (CO): Rockies
    # - Utah (UT): Rockies/Wasatch
    # - California (CA): Tahoe/Mammoth
    # - Vermont (VT): New England
    # - New Hampshire (NH): New England
    # - Washington (WA): Pacific Northwest
    # - Wyoming (WY): Jackson Hole

    ski_states = [
        ('US', 'U.S. Average'),
        ('CO', 'Colorado'),
        ('UT', 'Utah'),
        ('CA', 'California'),
        ('VT', 'Vermont'),
        ('NH', 'New Hampshire'),
        ('WA', 'Washington'),
        ('WY', 'Wyoming'),
    ]

    for state_id, state_name in ski_states:
        print_safe(f"\nFetching {state_name} Electricity (Commercial)...")
        # Use commercial sector (COM) - most relevant for ski operations
        elec_data = fetch_eia_electricity(state_id, sector='COM', limit=24)
        if elec_data:
            key = f'electricity_{state_id.lower()}'
            dashboard_data[key] = elec_data
            # EIA returns cents/kWh, display as such
            print_safe(f"  OK {state_id}: {elec_data[-1]['value']:.2f} cents/kWh ({len(elec_data)} data points)")
        else:
            print_safe(f"  ! {state_id}: No data available")

    # 10. 10-Year Treasury Yield (Daily - fetch ~1 year of trading days)
    print_safe("\nFetching 10-Year Treasury Yield...")
    treasury_data = fetch_fred_data('DGS10', limit=260)
    if treasury_data:
        dashboard_data['treasury_10y'] = treasury_data
        print_safe(f"  OK Latest: {treasury_data[-1]['value']}% ({len(treasury_data)} data points)")

    # 10b. 2-Year Treasury Yield (Daily - for yield curve spread)
    print_safe("\nFetching 2-Year Treasury Yield...")
    treasury_2y = fetch_fred_data('DGS2', limit=260)
    if treasury_2y:
        dashboard_data['treasury_2y'] = treasury_2y
        print_safe(f"  OK Latest: {treasury_2y[-1]['value']}% ({len(treasury_2y)} data points)")

    # 10c. Yield Curve Spread (10Y - 2Y) - computed from real data
    if treasury_data and treasury_2y:
        t10_dict = {d['date']: d['value'] for d in treasury_data}
        spread_data = []
        for d in treasury_2y:
            if d['date'] in t10_dict:
                spread_data.append({
                    'date': d['date'],
                    'value': round(t10_dict[d['date']] - d['value'], 2)
                })
        if spread_data:
            dashboard_data['yield_curve_spread'] = spread_data
            print_safe(f"  OK Yield Curve Spread: {spread_data[-1]['value']:+.2f}% ({len(spread_data)} data points)")

    # 11. Currency Exchange Rates (Daily - fetch ~1 year of trading days)
    # All rates stored as "units of foreign currency per 1 USD" so UP = stronger USD
    print_safe("\nFetching Currency Exchange Rates...")

    # USD/CAD (Canadian Dollars per USD) - DEXCAUS is CAD per USD, so up = stronger USD
    usd_cad = fetch_fred_data('DEXCAUS', limit=260)
    if usd_cad:
        dashboard_data['usd_cad'] = usd_cad
        print_safe(f"  OK USD/CAD: {usd_cad[-1]['value']} ({len(usd_cad)} data points)")

    # EUR/USD - DEXUSEU is USD per EUR, need to INVERT so up = stronger USD
    eur_usd_raw = fetch_fred_data('DEXUSEU', limit=260)
    if eur_usd_raw:
        # Invert: store as EUR per USD (1/rate)
        usd_eur = [{
            'date': d['date'],
            'value': round(1.0 / d['value'], 4) if d['value'] != 0 else 0
        } for d in eur_usd_raw]
        dashboard_data['usd_eur'] = usd_eur
        print_safe(f"  OK USD/EUR: {usd_eur[-1]['value']} ({len(usd_eur)} data points) [inverted]")

    # USD/JPY (Yen per USD) - DEXJPUS is JPY per USD, so up = stronger USD
    usd_jpy = fetch_fred_data('DEXJPUS', limit=260)
    if usd_jpy:
        dashboard_data['usd_jpy'] = usd_jpy
        print_safe(f"  OK USD/JPY: {usd_jpy[-1]['value']} ({len(usd_jpy)} data points)")

    # USD/MXN (Pesos per USD) - DEXMXUS is MXN per USD, so up = stronger USD
    usd_mxn = fetch_fred_data('DEXMXUS', limit=260)
    if usd_mxn:
        dashboard_data['usd_mxn'] = usd_mxn
        print_safe(f"  OK USD/MXN: {usd_mxn[-1]['value']} ({len(usd_mxn)} data points)")

    # USD/GBP - DEXUSUK is USD per GBP, need to INVERT so up = stronger USD
    gbp_usd_raw = fetch_fred_data('DEXUSUK', limit=260)
    if gbp_usd_raw:
        # Invert: store as GBP per USD (1/rate)
        usd_gbp = [{
            'date': d['date'],
            'value': round(1.0 / d['value'], 4) if d['value'] != 0 else 0
        } for d in gbp_usd_raw]
        dashboard_data['usd_gbp'] = usd_gbp
        print_safe(f"  OK USD/GBP: {usd_gbp[-1]['value']} ({len(usd_gbp)} data points) [inverted]")

    # USD/AUD - DEXUSAL is USD per AUD, need to INVERT so up = stronger USD
    aud_usd_raw = fetch_fred_data('DEXUSAL', limit=260)
    if aud_usd_raw:
        # Invert: store as AUD per USD (1/rate)
        usd_aud = [{
            'date': d['date'],
            'value': round(1.0 / d['value'], 4) if d['value'] != 0 else 0
        } for d in aud_usd_raw]
        dashboard_data['usd_aud'] = usd_aud
        print_safe(f"  OK USD/AUD: {usd_aud[-1]['value']} ({len(usd_aud)} data points) [inverted]")

    # USD/CNY (Yuan per USD) - DEXCHUS is CNY per USD, so up = stronger USD
    usd_cny = fetch_fred_data('DEXCHUS', limit=260)
    if usd_cny:
        dashboard_data['usd_cny'] = usd_cny
        print_safe(f"  OK USD/CNY: {usd_cny[-1]['value']} ({len(usd_cny)} data points)")

    # USD/INR (Rupees per USD) - DEXINUS is INR per USD, so up = stronger USD
    usd_inr = fetch_fred_data('DEXINUS', limit=260)
    if usd_inr:
        dashboard_data['usd_inr'] = usd_inr
        print_safe(f"  OK USD/INR: {usd_inr[-1]['value']} ({len(usd_inr)} data points)")

    # USD/KRW (Korean Won per USD) - Daily FRED series
    # South Korea is a growing luxury ski market with direct flights to Western US
    usd_krw = fetch_fred_data('DEXKOUS', limit=260)
    if usd_krw:
        dashboard_data['usd_krw'] = usd_krw
        print_safe(f"  OK USD/KRW: {usd_krw[-1]['value']} ({len(usd_krw)} data points)")

    # --- LUXURY FEEDER MARKET CURRENCIES ---
    print_safe("\nFetching Luxury Feeder Market Currencies...")

    # USD/HKD (Hong Kong Dollars per USD) - DEXHKUS
    usd_hkd = fetch_fred_data('DEXHKUS', limit=260)
    if usd_hkd:
        dashboard_data['usd_hkd'] = usd_hkd
        print_safe(f"  OK USD/HKD: {usd_hkd[-1]['value']} ({len(usd_hkd)} data points)")

    # USD/SGD (Singapore Dollars per USD) - DEXSIUS
    usd_sgd = fetch_fred_data('DEXSIUS', limit=260)
    if usd_sgd:
        dashboard_data['usd_sgd'] = usd_sgd
        print_safe(f"  OK USD/SGD: {usd_sgd[-1]['value']} ({len(usd_sgd)} data points)")

    # USD/CHF (Swiss Francs per USD) - DEXSZUS - for competitive destination comparison
    usd_chf = fetch_fred_data('DEXSZUS', limit=260)
    if usd_chf:
        dashboard_data['usd_chf'] = usd_chf
        print_safe(f"  OK USD/CHF: {usd_chf[-1]['value']} ({len(usd_chf)} data points)")

    # USD/BRL (Brazilian Reals per USD) - DEXBZUS - Aspen/Vail South American market
    usd_brl = fetch_fred_data('DEXBZUS', limit=260)
    if usd_brl:
        dashboard_data['usd_brl'] = usd_brl
        print_safe(f"  OK USD/BRL: {usd_brl[-1]['value']} ({len(usd_brl)} data points)")

    # USD/AED (UAE Dirham) - Pegged at 3.6725, store static value with latest date
    # No FRED daily series; AED has been pegged to USD since 1997
    if usd_cad:  # Use CAD dates as reference
        dashboard_data['usd_aed'] = [{'date': d['date'], 'value': 3.6725} for d in usd_cad[-30:]]
        print_safe(f"  OK USD/AED: 3.6725 (pegged) ({len(dashboard_data['usd_aed'])} data points)")

    # Fed Trade-Weighted Dollar Index (Broad) - DTWEXBGS
    # Measures USD against 26 major trading partners, weighted by trade volume
    # More representative than DXY for actual trade relationships
    print_safe("\nFetching Dollar Strength Indices...")
    trade_weighted = fetch_fred_data('DTWEXBGS', limit=260)
    if trade_weighted:
        dashboard_data['trade_weighted_usd'] = trade_weighted
        print_safe(f"  OK Trade-Weighted USD: {trade_weighted[-1]['value']} ({len(trade_weighted)} data points)")

    # Real Trade-Weighted Dollar Index (inflation-adjusted)
    real_trade_weighted = fetch_fred_data('RTWEXBGS', limit=260)
    if real_trade_weighted:
        dashboard_data['real_trade_weighted_usd'] = real_trade_weighted
        print_safe(f"  OK Real Trade-Weighted USD: {real_trade_weighted[-1]['value']} ({len(real_trade_weighted)} data points)")

    # 12. Commodities (Daily prices)
    print_safe("\nFetching Commodity Prices...")

    # Gold Price (USD per troy ounce) from FreeGoldAPI.com
    # FRED discontinued LBMA gold prices in Jan 2022, using free alternative API
    gold_data = fetch_gold_price()
    if gold_data:
        dashboard_data['gold'] = gold_data
        print_safe(f"  OK Gold: ${gold_data[-1]['value']:.2f}/oz ({len(gold_data)} data points)")
    else:
        print_safe("  ! Gold data unavailable")

    # Crude Oil WTI (USD per barrel)
    oil_data = fetch_fred_data('DCOILWTICO', limit=260)
    if oil_data:
        dashboard_data['crude_oil'] = oil_data
        print_safe(f"  OK Crude Oil WTI: ${oil_data[-1]['value']:.2f}/bbl ({len(oil_data)} data points)")

    # Natural Gas Henry Hub (USD per MMBtu)
    natgas_data = fetch_fred_data('DHHNGSP', limit=260)
    if natgas_data:
        dashboard_data['natural_gas'] = natgas_data
        print_safe(f"  OK Natural Gas: ${natgas_data[-1]['value']:.2f}/MMBtu ({len(natgas_data)} data points)")

    # Copper Price - FRED PCOPPUSDM is USD per metric ton, convert to USD per pound
    # 1 metric ton = 2204.62 pounds
    copper_data_raw = fetch_fred_data('PCOPPUSDM', limit=12)  # Monthly data
    if copper_data_raw:
        # Convert from USD/metric ton to USD/lb
        copper_data = [{
            'date': d['date'],
            'value': round(d['value'] / 2204.62, 2)
        } for d in copper_data_raw]
        dashboard_data['copper'] = copper_data
        print_safe(f"  OK Copper: ${copper_data[-1]['value']:.2f}/lb ({len(copper_data)} data points)")

    # 13. Daily News Sentiment Index (from SF Fed - updates weekly)
    print_safe("\nFetching News Sentiment Index...")
    news_sentiment = fetch_news_sentiment()
    if news_sentiment:
        dashboard_data['news_sentiment'] = news_sentiment
        print_safe(f"  OK Latest: {news_sentiment[-1]['value']:.3f} ({news_sentiment[-1]['date']}) ({len(news_sentiment)} data points)")

    # --- US/CANADA BORDER CROSSINGS (BTS) ---
    # Key indicator for Canadian visitor flows to US ski markets
    print_safe("\n--- Border Crossings (BTS US-Canada) ---")
    print_safe("\nFetching US/Canada Border Crossing Data...")
    border_data = fetch_bts_border_crossings(limit=36)
    if border_data:
        # National totals with 2019 baseline
        if border_data.get('national'):
            dashboard_data['border_national'] = border_data['national']
            dashboard_data['border_national_baseline_2019'] = border_data.get('national_baseline_2019')
            dashboard_data['border_national_vs_2019'] = border_data.get('national_vs_2019')
            latest = border_data['national'][-1]
            vs_2019 = border_data.get('national_vs_2019')
            vs_2019_str = f" ({vs_2019:+.1f}% vs 2019)" if vs_2019 else ""
            print_safe(f"  OK National: {latest['value']:,} passengers{vs_2019_str}")

        # State-level data for ski-relevant states with baselines
        if border_data.get('states'):
            for state_key, state_data in border_data['states'].items():
                dashboard_data[f'border_{state_key}'] = state_data
            dashboard_data['border_state_baselines'] = border_data.get('state_baselines', {})
            print_safe(f"  OK States: {', '.join(border_data['states'].keys())} ({len(border_data['states'])} states)")

        # Port-level data with resort mappings and 2019 baselines
        if border_data.get('ports'):
            dashboard_data['border_ports'] = border_data['ports']
            for port_key, port_info in border_data['ports'].items():
                dashboard_data[f'border_port_{port_key}'] = port_info.get('data', [])
            print_safe(f"  OK Ports: {', '.join(border_data['ports'].keys())} ({len(border_data['ports'])} ports)")
    else:
        print_safe("  ! Border crossing data unavailable")

    # --- CANADIAN OUTBOUND TRAVEL (Statistics Canada) ---
    # Canadian perspective: which provinces are sending travelers to US
    print_safe("\nFetching Canadian Outbound Travel (StatCan)...")
    statcan_outbound = fetch_statcan_canadian_outbound()
    if statcan_outbound:
        dashboard_data['canadian_outbound'] = statcan_outbound
        canada_total = statcan_outbound.get('canada_total', {})
        vs_2019 = canada_total.get('vs_2019')
        vs_2019_str = f" ({vs_2019:+.1f}% vs 2019)" if vs_2019 else ""
        print_safe(f"  OK Canada Total: {canada_total.get('latest', 0):,}K visits{vs_2019_str}")
        for key in ['alberta', 'british_columbia', 'ontario', 'quebec']:
            prov = statcan_outbound.get(key, {})
            if prov:
                print_safe(f"      {prov['name']}: {prov.get('latest', 0):,}K -> {prov.get('feeder_for', '')}")
    else:
        print_safe("  ! StatCan outbound data unavailable")

    # --- AIR TRAVEL INDICATORS ---
    print_safe("\n--- Air Travel Indicators ---")

    # Jet Fuel Prices (FRED weekly data) - leading indicator for airfares
    print_safe("\nFetching Jet Fuel Prices (Gulf Coast)...")
    jet_fuel = fetch_fred_data('WJFUELUSGULF', limit=104)  # 2 years of weekly data
    if jet_fuel:
        dashboard_data['jet_fuel'] = jet_fuel
        print_safe(f"  OK Jet Fuel: ${jet_fuel[-1]['value']:.2f}/gal ({len(jet_fuel)} weeks)")

    # National Enplanements (FRED monthly) - overall air travel demand
    print_safe("\nFetching National Enplanements...")
    enplanements = fetch_fred_data('ENPLANED11', limit=36)
    if enplanements:
        dashboard_data['enplanements'] = enplanements
        latest = enplanements[-1]['value']
        print_safe(f"  OK Enplanements: {latest:,.0f} ({len(enplanements)} months)")

    # Domestic Load Factor (FRED monthly) - capacity utilization
    print_safe("\nFetching Domestic Load Factor...")
    load_factor = fetch_fred_data('LOADFACTORD11', limit=36)
    if load_factor:
        dashboard_data['load_factor'] = load_factor
        print_safe(f"  OK Load Factor: {load_factor[-1]['value']:.1f}% ({len(load_factor)} months)")

    # TSA Checkpoint Data (daily) - real-time travel activity
    print_safe("\nFetching TSA Checkpoint Data...")
    tsa_data = fetch_tsa_checkpoint_data(limit=90)  # Last 90 days
    if tsa_data:
        dashboard_data['tsa_checkpoint'] = tsa_data
        latest = tsa_data[-1]
        print_safe(f"  OK TSA: {latest['value']:,} travelers ({latest['date']}) ({len(tsa_data)} days)")

    # T-100 Aggregate Passengers (monthly) - airline-reported traffic
    print_safe("\nFetching T-100 Aggregate Passengers...")
    t100_data = fetch_bts_t100_aggregate(limit=36)
    if t100_data:
        dashboard_data['t100_passengers'] = t100_data
        latest = t100_data[-1]
        print_safe(f"  OK T-100: {latest['value']:,} passengers ({latest['date'][:7]}) ({len(t100_data)} months)")

    print_safe("\nFetching Ski Gateway Airport Data...")
    ski_airports = fetch_ski_gateway_airports()
    if ski_airports:
        dashboard_data['ski_gateway_airports'] = ski_airports
        print_safe(f"  OK Ski Airports: {len(ski_airports)} airports loaded")
        # Show top 5 by passenger volume
        for apt in ski_airports[:5]:
            period = apt.get('month', apt.get('year', 'N/A'))
            print_safe(f"    {apt['code']} ({apt['city']}): {apt['passengers']:,} pax ({period})")

    # Build economic narrative from available data
    print_safe("\nBuilding economic narrative...")
    narrative = build_economic_narrative(dashboard_data)
    if narrative:
        dashboard_data['economic_narrative'] = narrative
        print_safe(f"  OK Narrative: {narrative[:80]}...")
    else:
        print_safe("  ! Insufficient data for narrative")

    # Write to file
    output_path = 'static/data/dashboard.json'
    print_safe(f"\nWriting data to {output_path}...")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with open(output_path, 'w') as f:
        json.dump(dashboard_data, f, indent=2)

    print_safe("\nOK Dashboard update complete!")
    print_safe(f"Updated {len([k for k in dashboard_data.keys() if k != 'updated'])} categories")

    return dashboard_data

if __name__ == '__main__':
    try:
        update_dashboard()
    except Exception as e:
        print_safe(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
