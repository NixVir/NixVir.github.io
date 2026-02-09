"""
Sports Betting Data Collection Script

Fetches U.S. sports betting handle/revenue data from Sports Handle
and Ontario data from iGaming Ontario, combines them into a unified dataset.

This serves as a "competing discretionary spend" indicator for the ski industry
economic dashboard.

Output: static/data/sports-betting.json
        static/data/sports-betting-history.json
"""

import io
import json
import re
import tempfile
import time
import requests
from datetime import datetime, date, timedelta, timezone
from pathlib import Path
from bs4 import BeautifulSoup

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Ontario Excel parsing disabled.")
    print("  Install with: pip install openpyxl")

# Output paths
OUTPUT_PATH = Path("static/data/sports-betting.json")
HISTORY_PATH = Path("static/data/sports-betting-history.json")

# CAD to USD conversion rate (approximate, as of early 2026)
CAD_TO_USD = 0.70

# U.S. adult population (18+), 2024 Census estimate
US_ADULT_POPULATION = 260_000_000

# Average ski lift ticket price (USD) for per-capita context
AVG_LIFT_TICKET_PRICE = 250

# ---------------------------------------------------------------------------
# BETTING EVENTS CALENDAR
# Dates snapped to 1st-of-month to align with the monthly x-axis on charts.
# type: milestone (one-time), recurring (recent annual), future (upcoming)
# tier: 1 = always show label, 2 = line only (visible on hover)
# ---------------------------------------------------------------------------
BETTING_EVENTS = [
    # Historical milestones
    {"date": "2018-05-01", "label": "PASPA Struck Down", "type": "milestone", "tier": 1,
     "category": "regulatory"},
    {"date": "2022-01-01", "label": "NY Mobile Launch", "type": "milestone", "tier": 1,
     "category": "regulatory"},

    # Recent recurring events (2025)
    {"date": "2025-02-01", "label": "Super Bowl LIX", "type": "recurring", "tier": 2,
     "category": "football"},
    {"date": "2025-03-01", "label": "March Madness '25", "type": "recurring", "tier": 2,
     "category": "basketball"},
    {"date": "2025-09-01", "label": "NFL Kickoff '25", "type": "recurring", "tier": 2,
     "category": "football"},

    # Future events (2026-2027)
    {"date": "2026-02-01", "label": "Super Bowl LX", "type": "future", "tier": 1,
     "category": "football"},
    {"date": "2026-03-01", "end_date": "2026-04-01", "label": "March Madness '26",
     "type": "future", "tier": 1, "category": "basketball"},
    {"date": "2026-05-01", "label": "Kentucky Derby", "type": "future", "tier": 2,
     "category": "horse-racing"},
    {"date": "2026-06-01", "end_date": "2026-07-01", "label": "FIFA World Cup 2026",
     "type": "future", "tier": 1, "category": "soccer"},
    {"date": "2026-09-01", "label": "NFL Kickoff '26", "type": "future", "tier": 1,
     "category": "football"},
    {"date": "2026-10-01", "label": "World Series '26", "type": "future", "tier": 2,
     "category": "baseball"},
    {"date": "2027-01-01", "label": "NFL Playoffs '27", "type": "future", "tier": 2,
     "category": "football"},
    {"date": "2027-02-01", "label": "Super Bowl LXI", "type": "future", "tier": 1,
     "category": "football"},
]


def fetch_sportshandle_data():
    """
    Scrape monthly U.S. sports betting data from Sports Handle.
    Returns list of monthly and annual records.
    """
    url = "https://sportshandle.com/sports-betting-revenue/"

    print("Fetching Sports Handle data...")

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }

    try:
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')

        # Find the main data table
        tables = soup.find_all('table')

        if not tables:
            print("  Warning: No tables found on page")
            return []

        # Look for the table with handle/revenue data
        data_table = None
        for table in tables:
            headers_row = table.find('tr')
            if headers_row:
                header_text = headers_row.get_text().lower()
                if 'handle' in header_text and 'revenue' in header_text:
                    data_table = table
                    break

        if not data_table:
            data_table = tables[0]
            print("  Using first table (no exact match found)")

        rows = data_table.find_all('tr')
        records = []

        def parse_currency(s):
            if not s:
                return None
            cleaned = re.sub(r'[,$]', '', s)
            try:
                return float(cleaned)
            except ValueError:
                return None

        for row in rows[1:]:  # Skip header row
            cells = row.find_all(['td', 'th'])
            if len(cells) < 4:
                continue

            cell_texts = [cell.get_text(strip=True) for cell in cells]

            # Skip empty rows
            if not cell_texts[0]:
                continue

            month_str = cell_texts[0]
            is_annual = 'TOTAL' in month_str.upper()

            # Parse dollar amounts from cells
            handle_str = None
            revenue_str = None
            taxes_str = None

            for i, text in enumerate(cell_texts[1:], 1):
                if '$' in text or (text and text[0].isdigit() and ',' in text):
                    cleaned = re.sub(r'[,$]', '', text)
                    try:
                        val = float(cleaned)
                        if val > 1_000_000_000:  # Handle (billions)
                            if handle_str is None:
                                handle_str = text
                        elif val > 100_000_000:  # Revenue or taxes
                            if revenue_str is None:
                                revenue_str = text
                            elif taxes_str is None:
                                taxes_str = text
                    except ValueError:
                        continue

            handle_usd = parse_currency(handle_str)
            revenue_usd = parse_currency(revenue_str)
            taxes_usd = parse_currency(taxes_str)

            # Parse date
            date_str = None
            if not is_annual and handle_usd:
                for fmt in ['%b-%y', '%B-%y', '%b %Y', '%B %Y', '%b-%Y']:
                    try:
                        parsed = datetime.strptime(month_str, fmt)
                        date_str = parsed.strftime('%Y-%m-01')
                        break
                    except ValueError:
                        continue

            if is_annual:
                year_match = re.search(r'(\d{4})', month_str)
                if year_match:
                    year = int(year_match.group(1))
                    records.append({
                        "type": "annual",
                        "year": year,
                        "handle_usd": handle_usd,
                        "revenue_usd": revenue_usd,
                        "taxes_usd": taxes_usd,
                        "source": "Sports Handle"
                    })
            elif date_str and handle_usd:
                records.append({
                    "type": "monthly",
                    "date": date_str,
                    "handle_usd": handle_usd,
                    "revenue_usd": revenue_usd,
                    "taxes_usd": taxes_usd,
                    "source": "Sports Handle"
                })

        monthly_count = len([r for r in records if r['type'] == 'monthly'])
        annual_count = len([r for r in records if r['type'] == 'annual'])
        print(f"  Found {monthly_count} monthly records")
        print(f"  Found {annual_count} annual summaries")

        return records

    except requests.exceptions.RequestException as e:
        print(f"  Error fetching Sports Handle: {e}")
        return []
    except Exception as e:
        print(f"  Error parsing Sports Handle: {e}")
        return []


def fetch_ontario_data():
    """
    Fetch Ontario iGaming sports betting data from Excel file.
    Returns list of monthly records with betting wagers and revenue.
    Falls back gracefully if openpyxl is unavailable or parsing fails.
    """
    print("Fetching Ontario iGaming data...")

    if not HAS_OPENPYXL:
        print("  Skipping: openpyxl not installed")
        return []

    page_url = "https://igamingontario.ca/en/operator/market-performance-report-monthly"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }

    try:
        response = requests.get(page_url, headers=headers, timeout=30)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')

        # Find Excel download link
        xlsx_link = None
        for link in soup.find_all('a', href=True):
            href = link['href']
            if '.xlsx' in href.lower():
                xlsx_link = href
                break

        if not xlsx_link:
            print("  Warning: No Excel download link found on page")
            # Try constructing URL for a recent month
            now = datetime.now()
            # Data lags ~6-8 weeks, try 2 months ago
            month_offset = now.month - 2
            if month_offset <= 0:
                target_year = now.year - 1
                target_month = month_offset + 12
            else:
                target_year = now.year
                target_month = month_offset
            target_date = datetime(target_year, target_month, 1)
            month_name = target_date.strftime('%B')
            xlsx_link = (
                f"https://igamingontario.ca/sites/default/files/documents/"
                f"iGO%20Monthly%20Market%20Performance%20Data%20Tables%20"
                f"-%20{target_year}%20{month_name}.xlsx"
            )

        if not xlsx_link.startswith('http'):
            xlsx_link = 'https://igamingontario.ca' + xlsx_link

        print(f"  Downloading: {xlsx_link}")

        xlsx_response = requests.get(xlsx_link, headers=headers, timeout=60)
        xlsx_response.raise_for_status()

        # Parse with openpyxl
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_response.content),
            data_only=True,
            read_only=True
        )

        print(f"  Sheets found: {wb.sheetnames}")

        records = []

        # Use "(Data) Product Monthly Stats" sheet which has per-product rows.
        # Structure (as of Dec 2025):
        #   Col 0: FiscalYearQuarter  (e.g., "FY22/23-Q1")
        #   Col 1: YearMonth          (e.g., "2022-04")
        #   Col 2: ProductCategory    (e.g., "BETTING", "CASINO", "P2P POKER")
        #   Col 3: CashWagers(M)      (number in millions CAD)
        #   Col 6: NAGGR(M)           (revenue in millions CAD)
        product_sheet_name = None
        for name in wb.sheetnames:
            if 'product' in name.lower() and 'data' in name.lower():
                product_sheet_name = name
                break

        if product_sheet_name:
            ws = wb[product_sheet_name]
            print(f"  Using sheet: '{product_sheet_name}'")

            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append(list(row))

            # Find the header row
            header_row_idx = None
            for idx, row in enumerate(all_rows):
                row_text = ' '.join(str(c or '').lower() for c in row)
                if 'yearmonth' in row_text or 'productcategory' in row_text:
                    header_row_idx = idx
                    break

            if header_row_idx is not None:
                header_cells = [str(c or '').strip().lower() for c in all_rows[header_row_idx]]
                print(f"  Header at row {header_row_idx + 1}: {header_cells}")

                # Map column indices
                yearmonth_col = None
                category_col = None
                wagers_col = None
                revenue_col = None

                for ci, cell in enumerate(header_cells):
                    if 'yearmonth' in cell:
                        yearmonth_col = ci
                    elif 'productcategory' in cell:
                        category_col = ci
                    elif 'cashwagers' in cell and 'mom' not in cell and 'market' not in cell:
                        wagers_col = ci
                    elif 'naggr' in cell and 'mom' not in cell and 'market' not in cell:
                        revenue_col = ci

                print(f"  Columns: yearmonth={yearmonth_col}, category={category_col}, "
                      f"wagers={wagers_col}, revenue={revenue_col}")

                if yearmonth_col is not None and category_col is not None and wagers_col is not None:
                    for row in all_rows[header_row_idx + 1:]:
                        if len(row) <= max(c for c in [yearmonth_col, category_col, wagers_col] if c is not None):
                            continue

                        category = str(row[category_col] or '').strip().upper()
                        if category != 'BETTING':
                            continue

                        yearmonth = str(row[yearmonth_col] or '').strip()
                        if not yearmonth or len(yearmonth) < 6:
                            continue

                        # Parse YearMonth "2022-04" -> "2022-04-01"
                        date_str = yearmonth + '-01' if len(yearmonth) == 7 else None

                        # Wagers in millions CAD
                        wager_val = row[wagers_col]
                        if wager_val is None:
                            continue
                        wager_m = float(wager_val) if isinstance(wager_val, (int, float)) else None
                        if wager_m is None:
                            continue
                        wager_cad = wager_m * 1_000_000  # Convert from millions

                        # Revenue (NAGGR) in millions CAD
                        rev_cad = None
                        if revenue_col is not None and revenue_col < len(row) and row[revenue_col] is not None:
                            rev_val = row[revenue_col]
                            if isinstance(rev_val, (int, float)):
                                rev_cad = float(rev_val) * 1_000_000

                        records.append({
                            "type": "monthly",
                            "date": date_str,
                            "region": "Ontario",
                            "betting_wagers_cad": wager_cad,
                            "betting_wagers_usd": wager_cad * CAD_TO_USD,
                            "betting_revenue_cad": rev_cad,
                            "betting_revenue_usd": rev_cad * CAD_TO_USD if rev_cad else None,
                            "source": "iGaming Ontario"
                        })
                else:
                    print("  Warning: Could not identify required columns")
            else:
                print("  Warning: Could not find header row in product data sheet")
        else:
            print("  Warning: No product data sheet found")

        wb.close()

        if records:
            print(f"  Parsed {len(records)} Ontario monthly records")
        else:
            print("  Warning: Could not parse monthly betting data from Excel")
            print("  The Excel file structure may have changed")

        return records

    except requests.exceptions.RequestException as e:
        print(f"  Error fetching Ontario data: {e}")
        return []
    except Exception as e:
        print(f"  Error processing Ontario data: {e}")
        import traceback
        traceback.print_exc()
        return []


def mark_incomplete(us_monthly):
    """
    Mark the most recent monthly record as incomplete if it lacks revenue data.
    This replaces hardcoded date exclusions in the frontend.
    """
    if not us_monthly:
        return

    # Sort descending to find most recent
    sorted_records = sorted(us_monthly, key=lambda r: r.get("date", ""), reverse=True)
    latest = sorted_records[0]

    if latest.get("revenue_usd") is None:
        latest["incomplete"] = True
        print(f"  Marked {latest.get('date')} as incomplete (missing revenue)")


def build_summary(us_data, ontario_data):
    """
    Build summary statistics for dashboard display.
    """
    summary = {
        "us": {},
        "ontario": {},
        "north_america": {}
    }

    # U.S. summary
    monthly_records = [r for r in us_data if r.get("type") == "monthly"]
    annual_records = [r for r in us_data if r.get("type") == "annual"]

    if monthly_records:
        monthly_records.sort(key=lambda x: x.get("date", ""), reverse=True)

        # Find latest non-incomplete month for display
        latest = monthly_records[0]
        summary["us"]["latest_month"] = latest.get("date")
        summary["us"]["latest_handle_usd"] = latest.get("handle_usd")
        summary["us"]["latest_revenue_usd"] = latest.get("revenue_usd")

        # Trailing 12 months (include all, even incomplete for handle sum)
        recent_12 = monthly_records[:12]
        ttm_handle = sum(r.get("handle_usd", 0) or 0 for r in recent_12)
        ttm_revenue = sum(r.get("revenue_usd", 0) or 0 for r in recent_12)
        summary["us"]["ttm_handle_usd"] = ttm_handle
        summary["us"]["ttm_revenue_usd"] = ttm_revenue

        # TTM YoY change
        if len(monthly_records) >= 24:
            prior_12 = monthly_records[12:24]
            prior_ttm_handle = sum(r.get("handle_usd", 0) or 0 for r in prior_12)
            if prior_ttm_handle > 0:
                ttm_yoy_change = (ttm_handle - prior_ttm_handle) / prior_ttm_handle
                summary["us"]["ttm_yoy_handle_change"] = ttm_yoy_change

        # Handle per capita
        if ttm_handle > 0:
            per_capita = ttm_handle / US_ADULT_POPULATION
            summary["us"]["handle_per_capita_usd"] = round(per_capita, 2)
            lift_ticket_equiv = per_capita / AVG_LIFT_TICKET_PRICE
            summary["us"]["handle_per_capita_context"] = f"~{lift_ticket_equiv:.0f} lift tickets"

    # Latest annual total
    if annual_records:
        annual_records.sort(key=lambda x: x.get("year", 0), reverse=True)
        latest_annual = annual_records[0]
        summary["us"]["latest_annual_year"] = latest_annual.get("year")
        summary["us"]["latest_annual_handle_usd"] = latest_annual.get("handle_usd")
        summary["us"]["latest_annual_revenue_usd"] = latest_annual.get("revenue_usd")

    # Ontario summary
    ontario_monthly = [r for r in ontario_data if r.get("type") == "monthly"]
    if ontario_monthly:
        # Compute Ontario TTM
        ontario_with_dates = [r for r in ontario_monthly if r.get("date")]
        if ontario_with_dates:
            ontario_with_dates.sort(key=lambda r: r["date"], reverse=True)
            ont_recent_12 = ontario_with_dates[:12]
            ont_ttm_wagers = sum(r.get("betting_wagers_usd", 0) or 0 for r in ont_recent_12)
            ont_ttm_revenue = sum(r.get("betting_revenue_usd", 0) or 0 for r in ont_recent_12)
            summary["ontario"]["ttm_wagers_usd"] = ont_ttm_wagers
            summary["ontario"]["ttm_revenue_usd"] = ont_ttm_revenue
            summary["ontario"]["latest_month"] = ont_recent_12[0].get("date")
            summary["ontario"]["months_available"] = len(ontario_with_dates)
        else:
            # Undated records — sum all wagers
            ont_total = sum(r.get("betting_wagers_usd", 0) or 0 for r in ontario_monthly)
            summary["ontario"]["total_wagers_usd"] = ont_total
            summary["ontario"]["months_available"] = len(ontario_monthly)
    elif ontario_data:
        # Fallback: non-monthly summary data
        first = ontario_data[0]
        summary["ontario"]["fiscal_year"] = first.get("fiscal_year")
        summary["ontario"]["betting_wagers_usd"] = first.get("betting_wagers_usd")
        summary["ontario"]["betting_revenue_usd"] = first.get("betting_revenue_usd")
        summary["ontario"]["note"] = first.get("note")

    # Payout reference line for chart
    summary["payout_reference"] = {
        "industry_avg_pct": 93.0,
        "label": "Typical bettor payout"
    }

    return summary


def build_narrative(summary, events):
    """
    Generate a plain-text interpretive narrative from current data.
    Follows CLAUDE.md language guidelines: factual, measured. Uses 'may' not 'will'.
    """
    parts = []

    ttm = summary.get("us", {}).get("ttm_handle_usd")
    yoy = summary.get("us", {}).get("ttm_yoy_handle_change")

    if ttm:
        ttm_b = ttm / 1e9
        parts.append(
            f"Since PASPA was struck down in May 2018, U.S. sports betting handle "
            f"has grown from $13.1B (2019) to ${ttm_b:.0f}B trailing twelve months"
        )

        if yoy is not None:
            yoy_pct = abs(yoy * 100)
            if yoy > 0.15:
                parts[-1] += (
                    f" \u2014 a {yoy_pct:.0f}% year-over-year increase. "
                    f"This accelerating growth in competing discretionary spending "
                    f"may be reducing the pool of potential new skiers, particularly "
                    f"among males 18-35."
                )
            elif yoy > 0:
                parts[-1] += (
                    f" \u2014 a {yoy_pct:.0f}% year-over-year increase. "
                    f"Growth is continuing but at a more moderate pace than "
                    f"the early post-PASPA expansion years."
                )
            else:
                parts[-1] += (
                    f" \u2014 a {yoy_pct:.0f}% year-over-year decline. "
                    f"A contraction in sports betting handle may free up "
                    f"discretionary income for other activities."
                )
        else:
            parts[-1] += "."

    # Per capita context
    per_capita = summary.get("us", {}).get("handle_per_capita_usd")
    if per_capita:
        parts.append(
            f"That works out to roughly ${per_capita:,.0f} per U.S. adult per year "
            f"\u2014 equivalent to about "
            f"{summary['us'].get('handle_per_capita_context', 'several lift tickets')}."
        )

    # Ontario context if available
    ont_ttm = summary.get("ontario", {}).get("ttm_wagers_usd")
    if ont_ttm and ont_ttm > 0:
        ont_b = ont_ttm / 1e9
        parts.append(
            f"Ontario, Canada's largest regulated market, has contributed "
            f"${ont_b:.1f}B USD in trailing twelve-month betting wagers."
        )

    # Next major future event
    future_events = [e for e in events if e.get("type") == "future" and e.get("tier") == 1]
    today_str = date.today().isoformat()
    upcoming = [e for e in future_events if e["date"] > today_str]
    if upcoming:
        upcoming.sort(key=lambda e: e["date"])
        next_event = upcoming[0]
        label = next_event["label"]

        if "FIFA" in label or "World Cup" in label:
            parts.append(
                f"The 2026 FIFA World Cup (June 11 \u2013 July 19) \u2014 the first "
                f"World Cup on U.S. soil with legal sports betting infrastructure "
                f"\u2014 may drive a significant handle spike."
            )
        elif "Super Bowl" in label:
            parts.append(
                f"The upcoming {label} may generate record single-event handle."
            )
        elif "March Madness" in label:
            parts.append(
                f"March Madness typically drives sustained handle volume "
                f"over three weeks of tournament play."
            )
        elif "NFL" in label:
            parts.append(
                f"The NFL season remains the single largest driver of "
                f"annual sports betting handle."
            )

    return " ".join(parts) if parts else None


def update_history(summary):
    """
    Append current snapshot to the history file.
    Keyed by latest_month (monthly cadence, not daily).
    Keeps 24 months. Returns previous snapshot for delta computation.
    """
    latest_month = summary.get("us", {}).get("latest_month")
    if not latest_month:
        return None

    # Load existing history
    if HISTORY_PATH.exists():
        with open(HISTORY_PATH, "r", encoding="utf-8") as f:
            history = json.load(f)
    else:
        history = {"snapshots": []}

    # Build snapshot
    snapshot = {
        "date": date.today().isoformat(),
        "latest_month": latest_month,
        "ttm_handle_usd": summary.get("us", {}).get("ttm_handle_usd"),
        "ttm_yoy_change": summary.get("us", {}).get("ttm_yoy_handle_change"),
        "latest_month_handle": summary.get("us", {}).get("latest_handle_usd"),
        "handle_per_capita_usd": summary.get("us", {}).get("handle_per_capita_usd"),
    }

    # Remove existing entry for this latest_month (idempotent reruns)
    history["snapshots"] = [
        s for s in history["snapshots"] if s.get("latest_month") != latest_month
    ]
    history["snapshots"].append(snapshot)

    # Keep last 24 entries, sorted by date
    history["snapshots"].sort(key=lambda s: s.get("date", ""))
    history["snapshots"] = history["snapshots"][-24:]

    # Find previous snapshot (the one before current)
    previous = None
    for s in history["snapshots"]:
        if s.get("latest_month") != latest_month:
            previous = s

    # Save
    HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(HISTORY_PATH, "w", encoding="utf-8") as f:
        json.dump(history, f, indent=2)

    return previous


def main():
    """Main entry point."""
    print("=" * 60)
    print("SPORTS BETTING DATA COLLECTION")
    print("=" * 60)
    print()

    # Fetch data from sources
    us_data = fetch_sportshandle_data()
    time.sleep(1)  # Be polite between requests
    ontario_data = fetch_ontario_data()

    # Mark incomplete records
    us_monthly = [r for r in us_data if r.get("type") == "monthly"]
    mark_incomplete(us_monthly)

    # Build summary
    summary = build_summary(us_data, ontario_data)

    # Build narrative
    narrative = build_narrative(summary, BETTING_EVENTS)
    if narrative:
        summary["narrative"] = narrative

    # Update history and get previous snapshot
    previous = update_history(summary)
    summary["previous"] = previous

    # Prepare output
    output = {
        "fetched_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "sources": {
            "us": "Sports Handle (sportshandle.com)",
            "ontario": "iGaming Ontario (igamingontario.ca)"
        },
        "us_monthly": [r for r in us_data if r.get("type") == "monthly"],
        "us_annual": [r for r in us_data if r.get("type") == "annual"],
        "ontario": ontario_data,
        "events": BETTING_EVENTS,
        "summary": summary
    }

    # Save to JSON
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(output, f, indent=2)

    print()
    print(f"Output saved to: {OUTPUT_PATH}")

    # Print summary
    print()
    print("SUMMARY:")
    print("-" * 40)

    if summary.get("us", {}).get("ttm_handle_usd"):
        ttm = summary["us"]["ttm_handle_usd"]
        print(f"U.S. Trailing 12-Month Handle: ${ttm/1e9:.1f}B")

    yoy = summary.get("us", {}).get("ttm_yoy_handle_change")
    if yoy is not None:
        print(f"U.S. YoY Handle Change: {yoy*100:+.1f}%")

    per_cap = summary.get("us", {}).get("handle_per_capita_usd")
    if per_cap:
        print(f"Handle Per Capita: ${per_cap:,.0f}/adult/year")
        print(f"  ({summary['us'].get('handle_per_capita_context', '')})")

    ont = summary.get("ontario", {})
    if ont.get("ttm_wagers_usd"):
        print(f"Ontario TTM Betting Wagers: ${ont['ttm_wagers_usd']/1e9:.1f}B USD")
    elif ont.get("betting_wagers_usd"):
        print(f"Ontario FY Betting Handle: ${ont['betting_wagers_usd']/1e9:.1f}B USD")

    if narrative:
        print()
        print("NARRATIVE:")
        print(narrative)

    if previous:
        print()
        print(f"PREVIOUS SNAPSHOT (for delta): {previous.get('date')}")
        print(f"  TTM Handle: ${previous.get('ttm_handle_usd', 0)/1e9:.1f}B")

    print()
    print(f"Events calendar: {len(BETTING_EVENTS)} events defined")
    print(f"History saved to: {HISTORY_PATH}")
    print()
    print("Done!")


if __name__ == "__main__":
    main()
